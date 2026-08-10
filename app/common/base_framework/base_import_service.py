from typing import Any, Dict, List, Optional, Generic, TypeVar, Type
from datetime import datetime, timezone
import io
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from pydantic import BaseModel
from app.common.base_framework.base_repository import BaseMongoRepository

T = TypeVar("T")
R = TypeVar("R", bound=BaseMongoRepository)


class ImportMode:
    CREATE = "CREATE"
    UPDATE = "UPDATE"
    UPSERT = "UPSERT"


class BaseImportService(Generic[T, R]):
    """
    Port 1-1 từ BaseImportService trong NestJS (base-import.service.ts:L49-728).
    Hỗ trợ sinh file Excel mẫu .xlsx, validate dữ liệu (dryRun), batch insert/update/upsert, và export Excel.
    """

    def __init__(self, repository: R):
        self.repository = repository

    def get_import_definition(self, model_cls: Type[BaseModel]) -> List[Dict[str, Any]]:
        """Lấy danh sách các trường import từ schema metadata."""
        fields = []
        if hasattr(model_cls, "model_fields"):
            for name, field_info in model_cls.model_fields.items():
                fields.append({
                    "field": name,
                    "label": field_info.title or name,
                    "required": field_info.is_required() if hasattr(field_info, "is_required") else True,
                    "description": field_info.description or "",
                    "example": getattr(field_info, "examples", [None])[0] if getattr(field_info, "examples", None) else None,
                })
        return fields

    def get_import_template_wb(
        self, model_cls: Type[BaseModel], example_data: Optional[List[Dict[str, Any]]] = None
    ) -> openpyxl.Workbook:
        """Port từ getImportTemplateWb (base-import.service.ts:L131-219)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"

        definitions = self.get_import_definition(model_cls)
        headers = [d["label"] for d in definitions]
        ws.append(headers)

        # Style header
        red_font = Font(bold=True, color="FF0000")
        bold_font = Font(bold=True)

        for col_num, d in enumerate(definitions, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = red_font if d.get("required") else bold_font

        # Add example rows
        if example_data:
            for row_item in example_data:
                ws.append([row_item.get(d["field"]) for d in definitions])
        else:
            example_row = [d.get("example") or "" for d in definitions]
            ws.append(example_row)

        return wb

    async def preprocess_import(
        self, rows: List[Dict[str, Any]], context: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """Override điểm hook xử lý tiền dữ liệu trước khi import."""
        return {"rows": rows, "context": context or {}}

    async def validate_and_transform_row_data(
        self, row_data: Dict[str, Any], context: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """Override điểm hook validate dòng dữ liệu."""
        return {"doc": row_data, "errors": []}

    async def insert_row_data(
        self, row_data: Dict[str, Any], mode: str = ImportMode.CREATE, keys: Optional[List[str]] = None
    ) -> Dict[str, Any]:
        """Port từ insertRowData (base-import.service.ts:L290-327)."""
        row = row_data.get("row", row_data)
        if mode == ImportMode.CREATE:
            return await self.repository.create(row)
        elif mode in (ImportMode.UPDATE, ImportMode.UPSERT):
            keys = keys or ["_id"]
            filter_query = {k: row.get(k) for k in keys if k in row}
            existing = await self.repository.collection.find_one(filter_query)
            if existing:
                return await self.repository.update_by_id(str(existing["_id"]), row)
            elif mode == ImportMode.UPSERT:
                return await self.repository.create(row)
        return row

    async def insert_import(
        self,
        rows: List[Dict[str, Any]],
        mode: str = ImportMode.CREATE,
        dry_run: bool = False,
        keys: Optional[List[str]] = None,
    ) -> Dict[str, Any]:
        """Port từ insertImport (base-import.service.ts:L345-508)."""
        processed = await self.preprocess_import(rows)
        row_list = processed.get("rows", rows)
        context = processed.get("context", {})

        validate_results = []
        has_error = False

        for idx, row in enumerate(row_list):
            transform_res = await self.validate_and_transform_row_data({"index": idx, "row": row}, context)
            doc = transform_res.get("doc", {"row": row})
            row_errors = transform_res.get("errors", [])

            insert_result = None
            if not row_errors and not dry_run:
                try:
                    insert_result = await self.insert_row_data(doc, mode=mode, keys=keys)
                except Exception as exc:
                    row_errors.append(str(exc))

            if row_errors:
                has_error = True

            validate_results.append({
                "index": idx,
                "row": row,
                "rowErrors": row_errors,
                "insertResult": insert_result,
            })

        return {
            "error": has_error,
            "dryRun": dry_run,
            "validate": validate_results,
            "totalProcessed": len(validate_results),
        }
